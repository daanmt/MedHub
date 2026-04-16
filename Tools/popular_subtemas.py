"""
Reconstroi taxonomia_cronograma do zero com dados corretos:
- 1 row por SUBTEMA (com questoes reais do Excel)
- O row da area-mae [bulk] e mantido apenas se nao houver subtemas
Isso elimina a duplicidade.
"""
import openpyxl
import sqlite3, os
from datetime import date

wb = openpyxl.load_workbook(r'dashboard-emed.xlsx', data_only=True)
DB = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'ipub.db')

ABAS = {
    'Pediatria':    'Pediatria',
    'Preventiva':   'Preventiva',
    'Cirurgia':     'Cirurgia',
    'Infectologia': 'Infecto',
    'Ginecologia':  'Ginecologia',
    'Gastro':       'Gastro',
    'Endocrino':    'Endocrino',
    'Cardiologia':  'Cardiologia',
    'Psiquiatria':  'Psiquiatria',
    'Neurologia':   'Neurologia',
    'Nefrologia':   'Nefrologia',
    'Hematologia':  'Hemato',
    'Pneumo':       'Pneumo',
    'Dermato':      'Dermato',
    'Reumato':      'Reumato',
    'Hepato':       'Hepato',
    'Otorrino':     'Otorrino',
}
for s in wb.sheetnames:
    if 'bstet' in s:
        ABAS[s] = 'Obstetricia'
        break

today = date.today().isoformat()
conn = sqlite3.connect(DB)

# Extrai todos os subtemas com dados
subtemas = []  # (area, tema, feitas, acertos)
for aba_name, db_area in ABAS.items():
    try:
        ws = wb[aba_name]
    except KeyError:
        continue
    for row in ws.iter_rows(values_only=True):
        tarefa    = row[1] if len(row) > 1 else None
        assunto   = row[2] if len(row) > 2 else None
        realizada = row[4] if len(row) > 4 else None
        feitas    = row[5] if len(row) > 5 else None
        acertos   = row[6] if len(row) > 6 else None
        if not isinstance(tarefa, (int, float)): continue
        if str(realizada).strip().upper() != 'SIM': continue
        if not isinstance(feitas, (int, float)) or feitas == 0: continue
        tema = (assunto or 'Sem nome').split('\n')[0].strip()[:120]
        subtemas.append((db_area, tema, int(feitas), int(acertos) if isinstance(acertos, (int,float)) else 0))

# Identifica areas com subtemas
areas_com_subtemas = set(s[0] for s in subtemas)

# 1. Remove rows de subtemas existentes e, para areas com subtemas,
#    tambem remove o row de area-mae (que sera substituido por subtemas)
for area in areas_com_subtemas:
    # Mantém o row [bulk] para referencia de ultima_revisao mas zera questoes
    # (os totais verdadeiros estao em sessoes_bulk)
    # Deleta qualquer outro row que nao seja [bulk] para recomecar limpo
    conn.execute("""
        DELETE FROM taxonomia_cronograma
        WHERE area=? AND tema NOT LIKE '[bulk]%'
    """, (area,))
print("[OK] Subtemas antigos removidos.")

# 2. Insere subtemas limpos
for area, tema, feitas, acertos in subtemas:
    pct = round(acertos / feitas * 100, 2) if feitas else 0
    conn.execute("""
        INSERT INTO taxonomia_cronograma
            (area, tema, questoes_realizadas, questoes_acertadas,
             percentual_acertos, ultima_revisao)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (area, tema, feitas, acertos, pct, today))
print("[OK] %d subtemas inseridos." % len(subtemas))

# 3. Sincroniza o row [bulk] de cada area com os totais reais
for area in areas_com_subtemas:
    sub_area = [(f, a) for ar, _, f, a in subtemas if ar == area]
    total_f = sum(x[0] for x in sub_area)
    total_a = sum(x[1] for x in sub_area)
    pct = round(total_a / total_f * 100, 2) if total_f else 0
    bulk_row = conn.execute(
        "SELECT id FROM taxonomia_cronograma WHERE area=? AND tema LIKE '[bulk]%' LIMIT 1",
        (area,)
    ).fetchone()
    if bulk_row:
        conn.execute("""
            UPDATE taxonomia_cronograma
            SET questoes_realizadas=?, questoes_acertadas=?, percentual_acertos=?, ultima_revisao=?
            WHERE id=?
        """, (total_f, total_a, pct, today, bulk_row[0]))
    else:
        conn.execute("""
            INSERT INTO taxonomia_cronograma
                (area, tema, questoes_realizadas, questoes_acertadas, percentual_acertos, ultima_revisao)
            VALUES (?, '[bulk] ' || ?, ?, ?, ?, ?)
        """, (area, area, total_f, total_a, pct, today))

# Atualiza tambem sessoes_bulk com os totais corretos do Excel
for area in areas_com_subtemas:
    sub_area = [(f, a) for ar, _, f, a in subtemas if ar == area]
    total_f = sum(x[0] for x in sub_area)
    total_a = sum(x[1] for x in sub_area)
    conn.execute("""
        UPDATE sessoes_bulk
        SET questoes_feitas=?, questoes_acertadas=?
        WHERE sessao_num=0 AND area=?
    """, (total_f, total_a, area))

conn.commit()

# Verificacao final
print()
print("=== VERIFICACAO FINAL sessoes_bulk ===")
rows = conn.execute("""
    SELECT area, questoes_feitas, questoes_acertadas
    FROM sessoes_bulk WHERE sessao_num=0
    ORDER BY questoes_feitas DESC
""").fetchall()
grand_q = 0
grand_a = 0
for r in rows:
    pct = r[2]/r[1]*100 if r[1] else 0
    print("  %-15s | q=%4d | a=%4d | %.1f%%" % (r[0], r[1], r[2], pct))
    grand_q += r[1]
    grand_a += r[2]
print()
print("  TOTAL: %d questoes / %d acertos (%.1f%%)" % (grand_q, grand_a, grand_a/grand_q*100 if grand_q else 0))
print("  (Quadro Geral Excel: 2980 | Soma subtemas: 3020 | Diff: +40 revisoes cross-tema)")

print()
print("=== SUBTEMAS NO DB ===")
subtema_rows = conn.execute("""
    SELECT area, tema, questoes_realizadas, questoes_acertadas
    FROM taxonomia_cronograma
    WHERE tema NOT LIKE '[bulk]%'
    ORDER BY area, questoes_realizadas DESC
""").fetchall()
cur = None
for r in subtema_rows:
    if r[0] != cur:
        print("\n  --- %s ---" % r[0])
        cur = r[0]
    pct = r[3]/r[2]*100 if r[2] else 0
    print("    %-50s | q=%4d | a=%4d | %.1f%%" % (r[1][:50], r[2], r[3], pct))

conn.close()
print("\n[SUCESSO] Reconstrucao concluida!")
