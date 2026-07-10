import os
import sys
import shutil
import tempfile
import unittest
import subprocess
from pathlib import Path

# Garante compatibilidade nativa de encoding em terminais Windows
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

ROOT_DIR = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(ROOT_DIR / "tools"))

import audit_resumos
import setup_hooks

try:
    import cronograma
    _CRONOGRAMA_OK = True
except Exception:
    _CRONOGRAMA_OK = False

try:
    import day_plan
    _DAY_PLAN_OK = True
except Exception:
    _DAY_PLAN_OK = False

try:
    import insert_questao
    _INSERT_OK = True
except Exception:
    _INSERT_OK = False

class TestAutonomiaHooks(unittest.TestCase):
    def setUp(self):
        self.test_resumo_ok = ROOT_DIR / "resumos" / "Cirurgia" / "Quadril Pediátrico.md"
        self.test_resumo_fail = ROOT_DIR / "resumos" / "_test_autonomia_fail.md"

    def tearDown(self):
        if self.test_resumo_fail.exists():
            try:
                self.test_resumo_fail.unlink()
            except Exception:
                pass

    def test_01_audit_resumos_pass_on_clean_file(self):
        """Verifica se audit_resumos aprova (exit code 0) um resumo canônico adequado."""
        if not self.test_resumo_ok.exists():
            self.skipTest("Arquivo de teste Quadril Pediátrico.md não encontrado.")
        
        exit_code = audit_resumos.audit_summaries([str(self.test_resumo_ok)])
        self.assertEqual(exit_code, 0, "O linter deveria aprovar o resumo limpo com exit code 0.")

    def test_02_audit_resumos_fail_on_invalid_file(self):
        """Verifica se audit_resumos reprova (exit code 1) um resumo com tabela e sem armadilhas."""
        content_fail = """# Resumo Inválido
| Tabela | Proibida |
| --- | --- |
| Erro | Erro |
Sem marcadores e sem secao de armadilhas.
"""
        with open(self.test_resumo_fail, "w", encoding="utf-8") as f:
            f.write(content_fail)

        exit_code = audit_resumos.audit_summaries([str(self.test_resumo_fail)])
        self.assertEqual(exit_code, 1, "O linter deveria reprovar o arquivo inválido com exit code 1.")

    def test_03_setup_hooks_lifecycle(self):
        """Ciclo de vida do instalador em um .git/hooks ISOLADO (tempdir).

        Hermético por construção: reatribui as constantes de path do módulo
        setup_hooks para um tempdir descartável e as restaura no finally. Nunca
        toca o .git/hooks real do repositório -- evitar isso é crítico porque o
        próprio pre-commit em execução estaria segurando esse arquivo (no Windows,
        um replace/unlink sobre ele dispararia sharing violation) e um teardown
        malfeito deixaria o hook do repo destruído ou armado indevidamente.
        """
        orig_git = setup_hooks.GIT_DIR
        orig_hooks = setup_hooks.HOOKS_DIR
        orig_pc = setup_hooks.PRE_COMMIT_PATH
        tmp = tempfile.mkdtemp(prefix="medhub_hooktest_")
        try:
            fake_hooks = Path(tmp) / ".git" / "hooks"
            fake_hooks.mkdir(parents=True)
            setup_hooks.GIT_DIR = Path(tmp) / ".git"
            setup_hooks.HOOKS_DIR = fake_hooks
            setup_hooks.PRE_COMMIT_PATH = fake_hooks / "pre-commit"

            # Instalar em ambiente limpo (sem hook prévio -> não gera .bak)
            self.assertTrue(setup_hooks.install(), "Falha na instalação do hook.")
            self.assertTrue(setup_hooks.PRE_COMMIT_PATH.exists(),
                            "O hook pre-commit deveria existir após a instalação.")
            content = setup_hooks.PRE_COMMIT_PATH.read_text(encoding="utf-8")
            self.assertIn("tools/auto_check.py --staged", content,
                          "O hook deveria invocar auto_check.py --staged.")

            # Reinstalar sobre hook existente -> gera backup .bak (idempotência)
            self.assertTrue(setup_hooks.install(), "Falha na reinstalação do hook.")
            self.assertTrue((fake_hooks / "pre-commit.bak").exists(),
                            "Reinstalar deveria preservar o anterior em pre-commit.bak.")

            # Uninstall -> restaura o backup por cima
            self.assertTrue(setup_hooks.uninstall(), "Falha no uninstall do hook.")
            self.assertTrue(setup_hooks.PRE_COMMIT_PATH.exists(),
                            "Uninstall deveria restaurar o hook a partir do .bak.")
        finally:
            setup_hooks.GIT_DIR = orig_git
            setup_hooks.HOOKS_DIR = orig_hooks
            setup_hooks.PRE_COMMIT_PATH = orig_pc
            shutil.rmtree(tmp, ignore_errors=True)

    def test_04_auto_check_cli_help(self):
        """Verifica se o orquestrador auto_check responde com sucesso na CLI."""
        cmd = [sys.executable, "-X", "utf8", "tools/auto_check.py", "--help"]
        res = subprocess.run(cmd, cwd=ROOT_DIR, capture_output=True, text=True)
        self.assertEqual(res.returncode, 0, "auto_check.py --help deveria retornar exit code 0.")
        self.assertIn("--changed", res.stdout)
        self.assertIn("--all", res.stdout)

    def test_05_diff_drive_captura_ordem(self):
        """Part 1: diff_drive anexa 'ordem' (linha do xlsx) por task; ordenar por
        'ordem' reflete a reordenacao manual do usuario, nao a sequencia do PDF."""
        if not _CRONOGRAMA_OK:
            self.skipTest("cronograma indisponivel")
        try:
            import openpyxl
            from openpyxl.styles import Font
        except Exception:
            self.skipTest("openpyxl indisponivel")
        tmp = tempfile.mkdtemp(prefix="medhub_ordem_")
        try:
            xlsx = Path(tmp) / "cron.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            # linha de datas: 10 colunas validas (parser exige >=10 semanas)
            for col in range(1, 11):
                ws.cell(row=cronograma.DRIVE_DATA_ROW, column=col, value="30/03 a 05/04")
            # semana 1 (col 1): Apendicite na linha 4, MFC na linha 5 (ordem do usuario)
            txt_ap = "Apendicite Aguda - Teoria - Link - 15 questoes"
            txt_mfc = "Medicina de Familia - Teoria - Link - 20 questoes"
            ws.cell(row=4, column=1, value=txt_ap)
            c_mfc = ws.cell(row=5, column=1, value=txt_mfc)
            c_mfc.font = Font(strike=True)      # MFC concluido (riscado)
            wb.save(xlsx)
            # grade stub: MFC ANTES de Apendicite (ordem do PDF) -- inverso do xlsx
            grade = {"semanas": [{"semana": 1, "tasks": [
                {"tarefa": "t_mfc", "area_norm": "MFC", "tema": "Medicina de Familia",
                 "tipo_norm": cronograma.normaliza_tipo(txt_mfc)},
                {"tarefa": "t_ap", "area_norm": "Cirurgia", "tema": "Apendicite Aguda",
                 "tipo_norm": cronograma.normaliza_tipo(txt_ap)},
            ]}]}
            res = cronograma.diff_drive(str(xlsx), grade=grade)
            by = {t["tarefa"]: t for t in res["tasks"]}
            self.assertEqual(by["t_ap"]["ordem"], 4, "Apendicite deveria herdar a linha 4 do xlsx.")
            self.assertEqual(by["t_mfc"]["ordem"], 5, "MFC deveria herdar a linha 5 do xlsx.")
            self.assertTrue(by["t_mfc"]["concluido"], "MFC riscado -> concluido True.")
            # ordenar por 'ordem' => Apendicite (4) antes de MFC (5): ordem do usuario, nao do PDF
            ordenado = sorted(res["tasks"], key=lambda t: t["ordem"])
            self.assertEqual([t["tarefa"] for t in ordenado], ["t_ap", "t_mfc"])
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_06_ordenar_por_drive_fallback(self):
        """Part 1: _ordenar_por_drive ordena por ordem_by_task (estavel, faltantes
        ao fim); sem ordem -> identidade (fallback PDF puro, DoD 3)."""
        if not _DAY_PLAN_OK:
            self.skipTest("day_plan indisponivel")
        tasks = [{"tarefa": "a"}, {"tarefa": "b"}, {"tarefa": "c"}]
        ordem = {(1, "a"): 5, (1, "b"): 4}      # 'c' sem ordem -> fim
        out = day_plan._ordenar_por_drive(tasks, ordem, 1)
        self.assertEqual([t["tarefa"] for t in out], ["b", "a", "c"])
        # sem ordem_by_task -> identidade (ordem do PDF preservada)
        self.assertEqual(day_plan._ordenar_por_drive(tasks, {}, 1), tasks)

    def test_07_material_efetivo_rebaixa_sem_md(self):
        """Part 2 (F30): 'resumo' vira 'extensivo' quando o tema nao tem .md;
        rotulo != 'resumo' passa direto; tema real com .md permanece 'resumo'."""
        if not _DAY_PLAN_OK:
            self.skipTest("day_plan indisponivel")
        # tema inexistente + rotulo 'resumo' -> rebaixa (nao promete "so ler o resumo")
        self.assertEqual(
            day_plan._material_efetivo("Tema Inexistente Xyz 999", "resumo"), "extensivo")
        # rotulo != 'resumo' -> passthrough (nao mexe)
        self.assertEqual(
            day_plan._material_efetivo("Tema Inexistente Xyz 999", "extensivo"), "extensivo")
        # tema real com .md -> permanece 'resumo'
        self.assertEqual(
            day_plan._material_efetivo("Quadril Pediátrico", "resumo"), "resumo")

    def test_08_tem_lastro_detecta_ausencia(self):
        """Part 2 (F31): _tem_lastro True p/ tema com .md; False p/ tema sem
        .md nem PDF-fonte par (candidato a Siamese Twins incompleto)."""
        if not _INSERT_OK:
            self.skipTest("insert_questao indisponivel")
        self.assertTrue(insert_questao._tem_lastro("Quadril Pediátrico"))
        self.assertFalse(insert_questao._tem_lastro("Tema Inexistente Xyz 999"))

    def test_09_contador_resumos_bate_com_linter(self):
        """Part 3 (higiene): _contar_resumos() usa o MESMO glob do audit_resumos
        -> numero identico ao do linter (fim do 63x61 digitado a mao no ESTADO)."""
        if not _DAY_PLAN_OK:
            self.skipTest("day_plan indisponivel")
        import glob as _glob
        esperado = len(_glob.glob(str(audit_resumos.TEMAS_DIR / "**" / "*.md"), recursive=True))
        self.assertEqual(day_plan._contar_resumos(), esperado)
        self.assertGreater(day_plan._contar_resumos(), 0)

if __name__ == "__main__":
    unittest.main(verbosity=2)
