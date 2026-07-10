"""test_orquestrador.py -- testes da part-2 (recomendador do dia, PRD orquestracao).

Fixtures DETERMINISTICAS sobre a funcao pura recomendar_dia() + queries novas de
db (db temp) + paridade contrato<->CLI. Pytest-nativo (coleta direta); standalone:
python tools/test_orquestrador.py. Nada aqui toca o ipub.db real.
"""
import json
import os
import re
import sqlite3
import sys
import tempfile
from datetime import date, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import app.utils.db as db  # noqa: E402
import cronograma as cr    # noqa: E402
import day_plan as dp      # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SINAIS_BASE = {
    "dias_enamed": 70, "restante_grade_q": 4000, "ritmo_real": 50.0,
    "vencidos": 3, "teto_efetivo": 30, "backlog_novos": 361,
    "semana_conteudo": 13, "lag": 3, "tema_alvo": "Apendicite Aguda",
    "frescos_tema_alvo": [],
}


def _tipos(reco):
    return [b["tipo"] for b in reco["blocos"]]


def test_mix_normal():
    r = dp.recomendar_dia(dict(SINAIS_BASE))
    tipos = _tipos(r)
    assert "questoes" in tipos and "fsrs" in tipos, f"mix basico (got {tipos})"
    assert "descanso" not in tipos and "simulado" not in tipos, \
        f"semana 13 nao-multipla e energia default: sem slots especiais (got {tipos})"
    assert len(r["justificativa"]) >= 3, "justificativa cita >=3 sinais"
    assert r["defaults_assumidos"] is True, "sem tempo/energia -> defaults declarados"
    pj = r["projecao"]
    # ritmo 50 q/dia, 4000 restantes -> 80d para fechar; 70d disponiveis -> folga -10
    assert pj["dias_para_fechar"] == 80 and pj["folga_dias"] == -10, f"projecao (got {pj})"
    assert pj["ritmo_necessario"] == round(4000 / 70, 1), "ritmo necessario derivado"
    q = [b for b in r["blocos"] if b["tipo"] == "questoes"][0]
    assert q["qtd"] == int(dp.TEMPO_DEFAULT_H * dp.QUESTOES_POR_HORA *
                           dp.FATOR_ENERGIA[dp.ENERGIA_DEFAULT]), \
        "grade atrasada -> capacidade maxima do dia"


def test_descanso_energia_baixa_com_folga():
    s = dict(SINAIS_BASE, restante_grade_q=500, ritmo_real=40.0)  # 13d p/ fechar, folga 57
    r = dp.recomendar_dia(s, energia="baixa")
    tipos = _tipos(r)
    assert "descanso" in tipos, f"R2 dispara (got {tipos})"
    assert "questoes" not in tipos, "dia de descanso nao empilha tema novo"
    fsrs = [b for b in r["blocos"] if b["tipo"] == "fsrs"]
    assert fsrs and fsrs[0]["qtd"] <= dp.FSRS_LEVE_CAP, "FSRS leve capado"
    assert any("descanso" in j for j in r["justificativa"]), "justificativa nomeia o motivo"


def test_energia_baixa_sem_folga_nao_descansa():
    r = dp.recomendar_dia(dict(SINAIS_BASE), energia="baixa")  # folga -10
    assert "descanso" not in _tipos(r), "sem folga nao ha descanso mesmo com energia baixa"
    q = [b for b in r["blocos"] if b["tipo"] == "questoes"][0]
    esperado = int(dp.TEMPO_DEFAULT_H * dp.QUESTOES_POR_HORA * dp.FATOR_ENERGIA["baixa"])
    assert q["qtd"] == esperado, f"capacidade modulada pela energia (got {q['qtd']})"


def test_simulado_semana_multipla():
    s = dict(SINAIS_BASE, semana_conteudo=12)  # 12 % 4 == 0
    r = dp.recomendar_dia(s)
    assert "simulado" in _tipos(r), "R3: semana multipla dispara slot de simulado"


def test_tempo_curto_reduz_capacidade():
    r = dp.recomendar_dia(dict(SINAIS_BASE), tempo_h=1.0, energia="media")
    q = [b for b in r["blocos"] if b["tipo"] == "questoes"][0]
    assert q["qtd"] == int(1.0 * dp.QUESTOES_POR_HORA * dp.FATOR_ENERGIA["media"]), \
        f"tempo curto reduz o bloco (got {q['qtd']})"
    assert r["defaults_assumidos"] is False, "contexto informado -> sem defaults"


def test_frescos_viram_primeiro_bloco():
    s = dict(SINAIS_BASE, frescos_tema_alvo=[{"id": 730}, {"id": 729}])
    r = dp.recomendar_dia(s)
    assert r["blocos"][0]["tipo"] == "mini-drill", "R1: anti-reincidencia primeiro"
    assert r["blocos"][0]["qtd"] == 2, "qtd = numero de cards frescos"


def test_folga_positiva_limita_pelo_necessario():
    s = dict(SINAIS_BASE, restante_grade_q=700, ritmo_real=40.0)  # nec 10/dia; folga 52
    r = dp.recomendar_dia(s)
    q = [b for b in r["blocos"] if b["tipo"] == "questoes"][0]
    assert q["qtd"] == 10, f"sem atraso, alvo = ritmo necessario (got {q['qtd']})"


def test_paridade_contrato_cli():
    path = os.path.join(ROOT, "core", "contracts", "orquestracao-contract.md")
    with open(path, encoding="utf-8") as fh:
        texto = fh.read()
    params = re.findall(r"^\|\s*([A-Z_]+)\s*\|\s*([0-9.]+)\s*\|", texto, re.M)
    assert len(params) >= 7, f"tabela de parametros no contrato (got {len(params)})"
    for nome, valor in params:
        assert hasattr(dp, nome), f"constante {nome} do contrato ausente no CLI"
        v_cli = getattr(dp, nome)
        assert float(v_cli) == float(valor), \
            f"paridade {nome}: contrato={valor} CLI={v_cli}"


def test_render_inclui_recomendacao():
    p = {
        "data": "2026-07-06", "dormant": {"empty": True},
        "volume": {"total": 100, "acertos": 80, "hoje": 0, "mes": 0, "alvo_enamed": 12000,
                   "faltam": 11900, "dias_ate_marco": 70, "ritmo_alvo": 170.0},
        "fsrs": {"atrasados": 1, "hoje": 2, "backlog_novos": 10},
        "divida": {"atrasados": 1, "regime_divida": False, "teto_base": 30, "teto_efetivo": 30},
        "cronograma": None, "cronograma_hint": None, "sugestao_passo": "x",
        "recomendacao": dp.recomendar_dia(dict(SINAIS_BASE)),
    }
    out = dp.render(p)
    assert "Recomendação do dia" in out, "secao presente no render"
    assert "projeção" in out and "sinais:" in out, "projecao e sinais renderizados"
    assert "[defaults:" in out, "defaults declarados no output"


def _db_temp_minimo():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    con = sqlite3.connect(path)
    con.executescript("""
        CREATE TABLE sessoes_bulk (id INTEGER PRIMARY KEY AUTOINCREMENT, sessao_num INTEGER,
            area TEXT, questoes_feitas INTEGER, questoes_acertadas INTEGER,
            data_sessao DATE, observacoes TEXT);
        CREATE TABLE taxonomia_cronograma (id INTEGER PRIMARY KEY AUTOINCREMENT,
            area TEXT, tema TEXT);
        CREATE TABLE flashcards (id INTEGER PRIMARY KEY AUTOINCREMENT, questao_id INTEGER,
            tema_id INTEGER, tipo TEXT, frente_pergunta TEXT);
        CREATE TABLE fsrs_cards (card_id INTEGER PRIMARY KEY, state INTEGER, due TIMESTAMP);
    """)
    con.commit()
    con.close()
    return path


def test_get_ritmo_real_janela():
    tmp = _db_temp_minimo()
    orig = db.DB_PATH
    db.DB_PATH = tmp
    try:
        con = sqlite3.connect(tmp)
        con.execute("INSERT INTO sessoes_bulk (sessao_num, area, questoes_feitas, "
                    "questoes_acertadas, data_sessao) VALUES "
                    "(1, 'Cirurgia', 60, 44, date('now', '-1 day')),"
                    "(2, 'Pediatria', 40, 30, date('now', '-5 day')),"
                    "(3, 'Simulado', 90, 60, date('now', '-2 day')),"     # Simulado nao conta
                    "(4, 'Cirurgia', 100, 70, date('now', '-30 day'))")   # fora da janela
        con.commit()
        con.close()
        assert db.get_ritmo_real(14) == round(100 / 14, 1), "soma so a janela, sem Simulado"
    finally:
        db.DB_PATH = orig
        os.remove(tmp)


def test_get_fresh_error_cards():
    tmp = _db_temp_minimo()
    orig = db.DB_PATH
    db.DB_PATH = tmp
    try:
        con = sqlite3.connect(tmp)
        con.execute("INSERT INTO taxonomia_cronograma (id, area, tema) VALUES "
                    "(1, 'Cirurgia', 'Apendicite Aguda'), (2, 'Pediatria', 'Imunizacoes')")
        con.execute("INSERT INTO flashcards (id, tema_id, tipo, frente_pergunta) VALUES "
                    "(730, 1, 'elo_quebrado', 'caso classico...'),"
                    "(500, 1, 'elo_quebrado', 'antigo...'),"
                    "(731, 2, 'elo_quebrado', 'outro tema...')")
        # 730/731 frescos (due=agora); 500 velho; todos state 0
        con.execute("INSERT INTO fsrs_cards (card_id, state, due) VALUES "
                    "(730, 0, datetime('now')), "
                    "(500, 0, datetime('now', '-10 day')), "
                    "(731, 0, datetime('now'))")
        con.commit()
        con.close()
        frescos = db.get_fresh_error_cards(tema="Apendicite", janela_horas=48)
        ids = sorted(c["id"] for c in frescos)
        assert ids == [730], f"so o fresco DO TEMA entra (got {ids})"
        todos = db.get_fresh_error_cards(janela_horas=48)
        assert sorted(c["id"] for c in todos) == [730, 731], "sem filtro: frescos de todos os temas"
    finally:
        db.DB_PATH = orig
        os.remove(tmp)


# ───────────── Sync Drive / W8 (spec cronograma-sync-conclusao-drive) ─────────────

def _xlsx_temp(cells):
    """xlsx minimo em disco p/ teste: linha 2 com >=10 semanas 'DD/MM a DD/MM'
    (exigido por _parse_conclusao_xlsx), celulas conforme `cells` =
    [(semana, row, texto, strike), ...]."""
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan1"
    n_semanas = max(10, max((c[0] for c in cells), default=1))
    for col in range(1, n_semanas + 1):
        ws.cell(row=2, column=col, value=f"{col:02d}/01 a {(col + 6):02d}/01")
    for semana, row, texto, strike in cells:
        c = ws.cell(row=row, column=semana, value=texto)
        if strike:
            c.font = Font(strike=True)
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return path


def test_norm_tema_xlsx_ignora_acentos_e_quebras():
    a = cr._norm_tema_xlsx("Doenças\nInflamatória do Tecido\nConjuntivo II (Teoria)")
    b = cr._norm_tema_xlsx("doencas inflamatoria do tecido   conjuntivo ii (teoria)")
    assert a == b, f"normalizacao deve casar acentos/quebras/espacos (got {a!r} vs {b!r})"


def test_diff_drive_matching_por_semana_tema_tipo():
    xlsx = _xlsx_temp([
        (1, 4, "Apendicite Aguda (Teoria)", True),
        (1, 5, "Apendicite Aguda (Revisão)", False),
    ])
    try:
        grade = {"semanas": [{"semana": 1, "tasks": [
            {"tarefa": 1, "area_norm": "Cirurgia", "tema": "Apendicite Aguda", "tipo_norm": "teoria"},
            {"tarefa": 2, "area_norm": "Cirurgia", "tema": "Apendicite Aguda", "tipo_norm": "revisao"},
        ]}]}
        r = cr.diff_drive(xlsx, grade=grade)
        by_tarefa = {t["tarefa"]: t["concluido"] for t in r["tasks"]}
        assert by_tarefa[1] is True, "teoria riscada -> concluido"
        assert by_tarefa[2] is False, "revisao nao riscada -> pendente"
        assert not r["sem_match"], f"ambas tasks deveriam casar por (semana,tema,tipo) (got {r['sem_match']})"
    finally:
        os.remove(xlsx)


def test_diff_drive_sem_match_fica_pendente():
    xlsx = _xlsx_temp([(1, 4, "Outro Tema Qualquer (Teoria)", True)])
    try:
        grade = {"semanas": [{"semana": 1, "tasks": [
            {"tarefa": 1, "area_norm": "Cirurgia", "tema": "Apendicite Aguda", "tipo_norm": "teoria"},
        ]}]}
        r = cr.diff_drive(xlsx, grade=grade)
        assert r["tasks"][0]["concluido"] is False, "sem match no xlsx -> conservador (pendente)"
        assert len(r["sem_match"]) == 1, "tema sem match deve ser reportado (nao falha silente)"
    finally:
        os.remove(xlsx)


def test_parse_conclusao_xlsx_rejeita_estrutura_invalida():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan1"
    ws.cell(row=2, column=1, value="lixo nao-data")
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    try:
        try:
            cr._parse_conclusao_xlsx(path)
            assert False, "estrutura invalida deveria levantar ValueError"
        except ValueError:
            pass
    finally:
        os.remove(path)


def test_conclusao_drive_ausente():
    tmp = _db_temp_minimo()
    orig = db.DB_PATH
    db.DB_PATH = tmp
    try:
        drive = dp._conclusao_drive()
        assert drive["by_task"] is None and drive["fresco"] is False, "sem snapshot gravado -> ausente"
        assert drive["ordem_by_task"] == {} and drive["atualizado_em"] is None
    finally:
        db.DB_PATH = orig
        os.remove(tmp)


def test_conclusao_drive_desatualizada_cai_no_fallback():
    tmp = _db_temp_minimo()
    orig = db.DB_PATH
    db.DB_PATH = tmp
    try:
        snap = {"gerado_em": "2026-07-07T10:00:00",
                "tasks": [{"semana": 12, "tarefa": 1, "concluido": True}]}
        db.set_preparacao("cronograma_conclusao_drive", json.dumps(snap), fonte="drive_sync")
        ontem = (date.today() - timedelta(days=1)).isoformat() + "T10:00:00"
        con = sqlite3.connect(tmp)
        con.execute("UPDATE preparacao_estado SET atualizado_em = ? WHERE chave = ?",
                    (ontem, "cronograma_conclusao_drive"))
        con.commit()
        con.close()
        drive = dp._conclusao_drive()
        assert drive["fresco"] is False, "snapshot de ontem -> nao fresco (fallback pro calendario)"
        assert drive["by_task"] == {(12, 1): True}, "mapa ainda retorna (so a flag de frescor muda)"
    finally:
        db.DB_PATH = orig
        os.remove(tmp)


def test_conclusao_drive_fresca():
    tmp = _db_temp_minimo()
    orig = db.DB_PATH
    db.DB_PATH = tmp
    try:
        snap = {"gerado_em": "now", "tasks": [
            {"semana": 12, "tarefa": 1, "concluido": True},
            {"semana": 12, "tarefa": 2, "concluido": False}]}
        db.set_preparacao("cronograma_conclusao_drive", json.dumps(snap), fonte="drive_sync")
        drive = dp._conclusao_drive()
        assert drive["fresco"] is True, "snapshot de hoje -> fresco"
        assert drive["by_task"] == {(12, 1): True, (12, 2): False}
        assert drive["ordem_by_task"] == {}, "snapshot antigo sem 'ordem' -> mapa vazio (fallback PDF)"
    finally:
        db.DB_PATH = orig
        os.remove(tmp)


def test_cronograma_hoje_filtra_por_conclusao_fresca():
    """DoD 2/3: com snapshot fresco, task concluida some de 'temas'; snapshot velho
    cai no comportamento antigo (semana inteira) + sinaliza conclusao_desatualizada."""
    tmp = _db_temp_minimo()
    orig = db.DB_PATH
    db.DB_PATH = tmp
    grade = {"_meta": {}, "semanas": [{
        "semana": 12, "inicio": "2026-06-15", "fim": "2026-06-21",
        "total_questoes": 100, "n_tasks": 2,
        "tasks": [
            {"tarefa": 1, "area_norm": "Preventiva", "tema": "Tema Feito",
             "tipo_norm": "teoria", "material_indicado": "resumo"},
            {"tarefa": 2, "area_norm": "Preventiva", "tema": "Tema Pendente",
             "tipo_norm": "teoria", "material_indicado": "resumo"},
        ],
    }]}

    class _FakeCr:
        ENAMED = cr.ENAMED

        @staticmethod
        def load_grade():
            return grade

        @staticmethod
        def semana_corrente(g, hoje):
            return 12

        @staticmethod
        def get_semana(g, n):
            return g["semanas"][0] if n == 12 else None

    orig_mod = sys.modules.get("cronograma")
    sys.modules["cronograma"] = _FakeCr
    try:
        db.set_preparacao("semana_conteudo", 12, fonte="operador")
        snap = {"gerado_em": "now", "tasks": [{"semana": 12, "tarefa": 1, "concluido": True}]}
        db.set_preparacao("cronograma_conclusao_drive", json.dumps(snap), fonte="drive_sync")
        hoje = date.today()
        c = dp._cronograma_hoje(0, hoje)
        assert c["conclusao_desatualizada"] is False, "snapshot de hoje -> nao desatualizado"
        assert c["temas"] == ["Tema Pendente"], f"tema feito deve sumir (got {c['temas']})"

        ontem = (hoje - timedelta(days=1)).isoformat() + "T10:00:00"
        con = sqlite3.connect(tmp)
        con.execute("UPDATE preparacao_estado SET atualizado_em = ? WHERE chave = ?",
                    (ontem, "cronograma_conclusao_drive"))
        con.commit()
        con.close()
        c2 = dp._cronograma_hoje(0, hoje)
        assert c2["conclusao_desatualizada"] is True, "snapshot velho -> sinaliza desatualizado"
        assert set(c2["temas"]) == {"Tema Feito", "Tema Pendente"}, \
            f"fallback lista a semana inteira sem filtro (got {c2['temas']})"
    finally:
        if orig_mod is not None:
            sys.modules["cronograma"] = orig_mod
        else:
            sys.modules.pop("cronograma", None)
        db.DB_PATH = orig
        os.remove(tmp)


if __name__ == "__main__":
    fns = [test_mix_normal, test_descanso_energia_baixa_com_folga,
           test_energia_baixa_sem_folga_nao_descansa, test_simulado_semana_multipla,
           test_tempo_curto_reduz_capacidade, test_frescos_viram_primeiro_bloco,
           test_folga_positiva_limita_pelo_necessario, test_paridade_contrato_cli,
           test_render_inclui_recomendacao, test_get_ritmo_real_janela,
           test_get_fresh_error_cards,
           test_norm_tema_xlsx_ignora_acentos_e_quebras, test_diff_drive_matching_por_semana_tema_tipo,
           test_diff_drive_sem_match_fica_pendente, test_parse_conclusao_xlsx_rejeita_estrutura_invalida,
           test_conclusao_drive_ausente, test_conclusao_drive_desatualizada_cai_no_fallback,
           test_conclusao_drive_fresca, test_cronograma_hoje_filtra_por_conclusao_fresca]
    falhas = 0
    for fn in fns:
        try:
            fn()
            print("  OK  " + fn.__name__)
        except AssertionError as e:
            falhas += 1
            print("  XX  %s: %s" % (fn.__name__, e))
    print()
    if falhas:
        print("FALHOU: %d teste(s)" % falhas)
        sys.exit(1)
    print("TODOS OS TESTES PASSARAM (part-2)")
